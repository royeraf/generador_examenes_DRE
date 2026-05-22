"""
Script de importación de UGELes e Instituciones Educativas desde Excel.

Fuente: rptDatosInstitucionesEducativas_20260217102437.xlsx (DRE Huánuco)

Pasos:
  1. Limpia UGELes e IEs anteriores (conserva SIN-ASIGNAR id=1)
  2. Importa 11 UGELes reales vinculadas a sus provincias
  3. Importa 1875 IEs con código modular, nombre, ugel, distrito
  4. Importa InstitucionNivel (un nivel por IE)

Uso:
    cd backend
    venv/bin/python3 scripts/import_ugeles_ies.py ../rptDatosInstitucionesEducativas_20260217102437.xlsx
"""

import asyncio
import sys
from pathlib import Path
from dotenv import load_dotenv

# Cargar .env antes de importar módulos de la app
load_dotenv(Path(__file__).parent.parent / ".env")

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy import select
from app.core.database import AsyncSessionLocal
# Importar todos los modelos para que SQLAlchemy resuelva las relaciones
import app.models.usuario  # noqa: F401
import app.models.db_models  # noqa: F401
import app.models.ubigeo  # noqa: F401
from app.models.db_models import Ugel, InstitucionEducativa, InstitucionNivel
from app.models.ubigeo import Provincia, Distrito


UGEL_CODIGOS = {
    "UGEL Ambo":          "UGEL-AMBO",
    "UGEL Dos de Mayo":   "UGEL-DOS-DE-MAYO",
    "UGEL Huacaybamba":   "UGEL-HUACAYBAMBA",
    "UGEL Huamalíes":     "UGEL-HUAMALIES",
    "UGEL Huánuco":       "UGEL-HUANUCO",
    "UGEL Lauricocha":    "UGEL-LAURICOCHA",
    "UGEL Leoncio Prado": "UGEL-LEONCIO-PRADO",
    "UGEL Marañon":       "UGEL-MARANON",
    "UGEL Pachitea":      "UGEL-PACHITEA",
    "UGEL Puerto Inca":   "UGEL-PUERTO-INCA",
    "UGEL Yarowilca":     "UGEL-YAROWILCA",
}


def norm(s: str) -> str:
    return (s.strip().upper()
            .replace("Á","A").replace("É","E").replace("Í","I")
            .replace("Ó","O").replace("Ú","U").replace("Ñ","N"))


_NIVEL_MAP = {
    "A1": "inicial",   # Inicial - Cuna
    "A2": "inicial",   # Inicial - Jardín
    "A3": "inicial",   # Inicial - Cuna-Jardín
    "A5": "inicial",   # Inicial - Prog No Escolariz
    "B0": "primaria",  # Primaria
    "F0": "secundaria",# Secundaria
    "E1": "inicial",   # Básica Especial - Inicial
    "E2": "primaria",  # Básica Especial - Primaria
}

def parse_nivel(nivel_str: str) -> str:
    if not nivel_str:
        return ""
    codigo = nivel_str.strip().split(" ")[0].strip().upper()
    return _NIVEL_MAP.get(codigo, "")


async def limpiar_todo(db):
    # Redirigir usuarios a SIN-ASIGNAR
    from sqlalchemy import text
    await db.execute(text("UPDATE usuarios SET ugel_id = 1 WHERE ugel_id != 1 AND ugel_id IS NOT NULL"))
    await db.execute(text("UPDATE usuarios SET institucion_educativa_id = 1 WHERE institucion_educativa_id != 1 AND institucion_educativa_id IS NOT NULL"))
    # Eliminar codigos_clase vinculados
    await db.execute(text("DELETE FROM codigos_clase WHERE institucion_educativa_id != 1"))
    # Eliminar niveles
    await db.execute(text("DELETE FROM institucion_niveles WHERE institucion_id != 1"))
    # Eliminar IEs reales
    await db.execute(text("DELETE FROM instituciones_educativas WHERE id != 1"))
    # Eliminar UGELes reales
    await db.execute(text("DELETE FROM ugeles WHERE id != 1"))
    print("  Limpieza completa.")


async def importar_ugeles(db) -> dict[str, int]:
    # Construir mapa nombre_norm → provincia_id
    result = await db.execute(select(Provincia))
    provincias = {norm(p.nombre): p.id for p in result.scalars().all()}

    ugel_id_map: dict[str, int] = {}

    for nombre_excel, codigo in UGEL_CODIGOS.items():
        prov_key = norm(nombre_excel.replace("UGEL ", ""))
        provincia_id = provincias.get(prov_key)
        if not provincia_id:
            for k, v in provincias.items():
                if prov_key in k or k in prov_key:
                    provincia_id = v
                    break

        ugel = Ugel(codigo=codigo, nombre=nombre_excel, provincia_id=provincia_id, is_active=True)
        db.add(ugel)
        await db.flush()  # obtiene el id sin commit
        ugel_id_map[nombre_excel] = ugel.id
        print(f"  {nombre_excel} → id={ugel.id}")

    return ugel_id_map


async def importar_ies(db, data: list, ugel_id_map: dict[str, int]):
    result = await db.execute(select(Distrito))
    distritos = {norm(d.nombre): d.id for d in result.scalars().all()}

    creadas = 0
    sin_distrito = 0
    sin_ugel = 0

    for row in data:
        ugel_excel = str(row[1]).strip() if row[1] else ""
        dist_excel = str(row[4]).strip() if row[4] else ""
        cod_mod    = str(row[6]).strip() if row[6] else ""
        nombre_ie  = str(row[8]).strip() if row[8] else ""
        nivel_str  = str(row[9]).strip() if row[9] else ""

        if not cod_mod or not nombre_ie:
            continue

        ugel_id = ugel_id_map.get(ugel_excel)
        if not ugel_id:
            sin_ugel += 1
            ugel_id = 1

        distrito_id = distritos.get(norm(dist_excel))
        if not distrito_id:
            sin_distrito += 1

        ie = InstitucionEducativa(
            codigo_modular=cod_mod,
            nombre=nombre_ie,
            ugel_id=ugel_id,
            distrito_id=distrito_id,
            is_active=True,
        )
        db.add(ie)
        await db.flush()

        nivel_codigo = parse_nivel(nivel_str)
        if nivel_codigo:
            db.add(InstitucionNivel(institucion_id=ie.id, nivel=nivel_codigo))

        creadas += 1
        if creadas % 200 == 0:
            print(f"  ... {creadas} IEs procesadas")
            await db.flush()

    print(f"  Total IEs creadas: {creadas}")
    if sin_distrito:
        print(f"  Sin distrito en DB: {sin_distrito}")
    if sin_ugel:
        print(f"  Con UGEL desconocida: {sin_ugel}")


async def main(xlsx_path: str):
    print(f"Leyendo {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    data = [row for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True) if row[0] is not None]
    print(f"  {len(data)} registros leídos del Excel.")

    async with AsyncSessionLocal() as db:
        print("\n[1/3] Limpiando datos existentes...")
        await limpiar_todo(db)

        print("\n[2/3] Importando UGELes...")
        ugel_id_map = await importar_ugeles(db)

        print("\n[3/3] Importando Instituciones Educativas...")
        await importar_ies(db, data, ugel_id_map)

        print("\nGuardando en base de datos...")
        await db.commit()

    print("Importacion completada.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 scripts/import_ugeles_ies.py <ruta_excel>")
        sys.exit(1)
    asyncio.run(main(sys.argv[1]))

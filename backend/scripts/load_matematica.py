"""
Script para cargar las competencias, capacidades, estándares y desempeños de Matemática
desde los archivos Word en la base de datos SQLite.

Ejecutar desde el directorio backend:
    python -m scripts.load_matematica

Archivos fuente (en capacidades_mat/):
    - COMPETENCIA MATEMATICA 1.docx (Cantidad)
    - COMPETENCIA MATEMATICA 2.docx (Regularidad, equivalencia y cambio)
    - COMPETENCIA MATEMATICA 3.docx (Forma, movimiento y localización)
    - COMPETENCIA MATEMATICA 4.docx (Gestión de datos e incertidumbre)
"""

import sys
import os
import re
from docx import Document
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env'))

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from app.core.database import Base
from app.models.db_models import (
    Grado,
    CompetenciaMatematica,
    CapacidadMatematica,
    EstandarMatematica,
    DesempenoMatematica
)
# Import all models so Base.metadata.create_all() can resolve foreign keys
from app.models.docente import Docente  # noqa: F401
from app.models.ubigeo import Provincia, Distrito  # noqa: F401

# Motor síncrono para el script de carga
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./desempenos.db")
if DATABASE_URL.startswith("postgresql+asyncpg://"):
    DATABASE_URL = DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://")
elif DATABASE_URL.startswith("mysql+aiomysql://"):
    DATABASE_URL = DATABASE_URL.replace("mysql+aiomysql://", "mysql+pymysql://")

if DATABASE_URL.startswith("postgresql") or DATABASE_URL.startswith("mysql"):
    engine = create_engine(DATABASE_URL)
else:
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(bind=engine)


# Ruta a la carpeta con los archivos Word y Excel
DOCX_FOLDER = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "capacidades_mat"
)
EXCEL_DESEMPENOS = os.path.join(DOCX_FOLDER, "desempenos_extraidos.xlsx")

# Definición de las 4 competencias matemáticas
COMPETENCIAS = [
    {
        "codigo": 1,
        "nombre": "Resuelve problemas de cantidad",
        "descripcion": """Consiste en que el estudiante solucione problemas o plantee nuevos problemas que le demanden construir y comprender las nociones de cantidad, de número, de sistemas numéricos, sus operaciones y propiedades. Además, dotar de significado a estos conocimientos en la situación y usarlos para representar o reproducir las relaciones entre sus datos y condiciones. Implica también discernir si la solución buscada requiere darse como una estimación o cálculo exacto, y para ello selecciona estrategias, procedimientos, unidades de medida y diversos recursos.""",
        "capacidades": [
            "Traduce cantidades a expresiones numéricas",
            "Comunica su comprensión sobre los números y las operaciones",
            "Usa estrategias y procedimientos de estimación y cálculo",
            "Argumenta afirmaciones sobre las relaciones numéricas y las operaciones"
        ]
    },
    {
        "codigo": 2,
        "nombre": "Resuelve problemas de regularidad, equivalencia y cambio",
        "descripcion": """Consiste en que el estudiante logre caracterizar equivalencias y generalizar regularidades y el cambio de una magnitud con respecto de otra, a través de reglas generales que le permitan encontrar valores desconocidos, determinar restricciones y hacer predicciones sobre el comportamiento de un fenómeno.""",
        "capacidades": [
            "Traduce datos y condiciones a expresiones algebraicas y gráficas",
            "Comunica su comprensión sobre las relaciones algebraicas",
            "Usa estrategias y procedimientos para encontrar equivalencias y reglas generales",
            "Argumenta afirmaciones sobre relaciones de cambio y equivalencia"
        ]
    },
    {
        "codigo": 3,
        "nombre": "Resuelve problemas de forma, movimiento y localización",
        "descripcion": """Consiste en que el estudiante se oriente y describa la posición y el movimiento de objetos y de sí mismo en el espacio, visualizando, interpretando y relacionando las características de los objetos con formas geométricas bidimensionales y tridimensionales.""",
        "capacidades": [
            "Modela objetos con formas geométricas y sus transformaciones",
            "Comunica su comprensión sobre las formas y relaciones geométricas",
            "Usa estrategias y procedimientos para medir y orientarse en el espacio",
            "Argumenta afirmaciones sobre relaciones geométricas"
        ]
    },
    {
        "codigo": 4,
        "nombre": "Resuelve problemas de gestión de datos e incertidumbre",
        "descripcion": """Consiste en que el estudiante analice datos sobre un tema de interés o estudio o de situaciones aleatorias, que le permitan tomar decisiones, elaborar predicciones razonables y conclusiones respaldadas en la información producida.""",
        "capacidades": [
            "Representa datos con gráficos y medidas estadísticas o probabilísticas",
            "Comunica su comprensión de los conceptos estadísticos y probabilísticos",
            "Usa estrategias y procedimientos para recopilar y procesar datos",
            "Sustenta conclusiones o decisiones con base en la información obtenida"
        ]
    }
]

# Definición de grados (incluyendo Inicial)
GRADOS_MATEMATICA = [
    {"nombre": "INICIAL DE 5 AÑOS", "numero": 0, "nivel": "inicial", "orden": 1, "ciclo": "II"},
    {"nombre": "PRIMER GRADO DE PRIMARIA", "numero": 1, "nivel": "primaria", "orden": 2, "ciclo": "III"},
    {"nombre": "SEGUNDO GRADO DE PRIMARIA", "numero": 2, "nivel": "primaria", "orden": 3, "ciclo": "III"},
    {"nombre": "TERCER GRADO DE PRIMARIA", "numero": 3, "nivel": "primaria", "orden": 4, "ciclo": "IV"},
    {"nombre": "CUARTO GRADO DE PRIMARIA", "numero": 4, "nivel": "primaria", "orden": 5, "ciclo": "IV"},
    {"nombre": "QUINTO GRADO DE PRIMARIA", "numero": 5, "nivel": "primaria", "orden": 6, "ciclo": "V"},
    {"nombre": "SEXTO GRADO DE PRIMARIA", "numero": 6, "nivel": "primaria", "orden": 7, "ciclo": "V"},
    {"nombre": "PRIMER GRADO DE SECUNDARIA", "numero": 1, "nivel": "secundaria", "orden": 8, "ciclo": "VI"},
    {"nombre": "SEGUNDO GRADO DE SECUNDARIA", "numero": 2, "nivel": "secundaria", "orden": 9, "ciclo": "VI"},
    {"nombre": "TERCER GRADO DE SECUNDARIA", "numero": 3, "nivel": "secundaria", "orden": 10, "ciclo": "VII"},
    {"nombre": "CUARTO GRADO DE SECUNDARIA", "numero": 4, "nivel": "secundaria", "orden": 11, "ciclo": "VII"},
    {"nombre": "QUINTO GRADO DE SECUNDARIA", "numero": 5, "nivel": "secundaria", "orden": 12, "ciclo": "VII"},
]


def identify_grado(text: str) -> dict | None:
    """Identifica el grado a partir de un texto corto (título de sección)."""
    if len(text) > 80:
        return None
    
    text_upper = text.upper()
    
    for grado in GRADOS_MATEMATICA:
        if grado["nivel"] == "inicial":
            if "INICIAL" in text_upper and ("5 AÑOS" in text_upper or "5AÑOS" in text_upper):
                return grado
        elif "GRADO DE PRIMARIA" in text_upper or "GRADO DE SECUNDARIA" in text_upper:
            ordinals = {
                "PRIMER": 1, "SEGUNDO": 2, "TERCER": 3, "CUARTO": 4,
                "QUINTO": 5, "SEXTO": 6
            }
            for ordinal, num in ordinals.items():
                if ordinal in text_upper:
                    if "PRIMARIA" in text_upper and grado["nivel"] == "primaria" and grado["numero"] == num:
                        return grado
                    if "SECUNDARIA" in text_upper and grado["nivel"] == "secundaria" and grado["numero"] == num:
                        return grado
    
    return None


def is_paragraph_bold(para) -> bool:
    """Verifica si un párrafo tiene formato bold (total o mayoritariamente)."""
    if not para.runs:
        return False
    # Considerar bold si más del 50% del texto está en bold
    total_chars = 0
    bold_chars = 0
    for run in para.runs:
        text_len = len(run.text.strip())
        total_chars += text_len
        if run.bold:
            bold_chars += text_len
    if total_chars == 0:
        return False
    return (bold_chars / total_chars) > 0.5


def parse_docx_competencia(docx_path: str, competencia_codigo: int) -> dict:
    """
    Parsea un archivo .docx de competencia matemática usando python-docx.
    Usa el formato bold para distinguir entre títulos de capacidad y desempeños.
    Retorna un diccionario con estándares y desempeños por grado.
    """
    print(f"\n📄 Procesando: {os.path.basename(docx_path)}")
    
    try:
        doc = Document(docx_path)
    except Exception as e:
        print(f"   ⚠️ Error abriendo {docx_path}: {e}")
        return {"estandares": [], "desempenos": []}
    
    result = {
        "estandares": [],
        "desempenos": []
    }
    
    capacidades_competencia = COMPETENCIAS[competencia_codigo - 1]["capacidades"]
    
    current_grado = None
    current_capacidad_orden = None
    current_section = None  # "estandar", "desempenos"
    estandar_buffer = []
    
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        
        text_upper = text.upper()
        bold = is_paragraph_bold(para)
        
        # 1. Detectar cambio de grado (líneas cortas, generalmente bold)
        grado_detected = identify_grado(text)
        if grado_detected:
            # Guardar estándar anterior si existe
            if current_grado and estandar_buffer:
                result["estandares"].append({
                    "grado_info": current_grado,
                    "descripcion": ' '.join(estandar_buffer),
                    "ciclo": current_grado.get("ciclo", "")
                })
                estandar_buffer = []
            
            current_grado = grado_detected
            current_section = None
            current_capacidad_orden = None
            print(f"   📍 Grado: {current_grado['nombre']}")
            continue
        
        if not current_grado:
            continue
        
        # 2. Detectar sección de estándar
        if "ESTÁNDAR" in text_upper or "ESTANDAR" in text_upper:
            current_section = "estandar"
            estandar_buffer = []
            continue
        
        # 3. Detectar sección de desempeños
        if "DESEMPEÑO" in text_upper and "CAPACIDAD" in text_upper:
            if estandar_buffer:
                result["estandares"].append({
                    "grado_info": current_grado,
                    "descripcion": ' '.join(estandar_buffer),
                    "ciclo": current_grado.get("ciclo", "")
                })
                estandar_buffer = []
            current_section = "desempenos"
            continue
        
        # 4. Acumular contenido de estándar
        if current_section == "estandar" and not bold:
            if "COMPETENCIA" not in text_upper and "CAPACIDAD" not in text_upper:
                estandar_buffer.append(text)
            continue
        
        # 5. Procesar sección de desempeños
        if current_section == "desempenos":
            # Detectar formato primaria: párrafo mixto "CAPACIDAD bold: desempeño normal"
            has_bold_run = any(run.bold and run.text.strip() for run in para.runs)
            has_normal_run = any(not run.bold and run.text.strip() for run in para.runs)
            if has_bold_run and has_normal_run and not bold:
                bold_text = ''.join(run.text for run in para.runs if run.bold).strip().rstrip(':')
                normal_text = ''.join(run.text for run in para.runs if not run.bold).strip()
                # Intentar identificar la capacidad del texto bold
                for idx, cap_nombre in enumerate(capacidades_competencia):
                    cap_words = cap_nombre.upper().split()[:3]
                    if all(word in bold_text.upper() for word in cap_words):
                        current_capacidad_orden = idx + 1
                        print(f"      ✓ Capacidad {current_capacidad_orden}: {cap_nombre[:50]}...")
                        break
                if current_capacidad_orden and len(normal_text) > 20:
                    result["desempenos"].append({
                        "grado_info": current_grado,
                        "capacidad_orden": current_capacidad_orden,
                        "descripcion": normal_text
                    })
                continue

            # Si es bold → es un título de capacidad (formato secundaria)
            if bold:
                for idx, cap_nombre in enumerate(capacidades_competencia):
                    cap_words = cap_nombre.upper().split()[:3]
                    if all(word in text_upper for word in cap_words):
                        current_capacidad_orden = idx + 1
                        print(f"      ✓ Capacidad {current_capacidad_orden}: {cap_nombre[:50]}...")
                        break
                continue

            # Si NO es bold y tenemos capacidad actual → es un desempeño (formato secundaria)
            if current_capacidad_orden and len(text) > 20:
                result["desempenos"].append({
                    "grado_info": current_grado,
                    "capacidad_orden": current_capacidad_orden,
                    "descripcion": text
                })
    
    # Guardar último estándar si existe
    if current_grado and estandar_buffer:
        result["estandares"].append({
            "grado_info": current_grado,
            "descripcion": ' '.join(estandar_buffer),
            "ciclo": current_grado.get("ciclo", "")
        })
    
    print(f"   📊 Encontrados: {len(result['estandares'])} estándares, {len(result['desempenos'])} desempeños")
    
    return result


def load_desempenos_from_excel() -> dict:
    """
    Lee desempeños desde desempenos_extraidos.xlsx (hojas C1-C4).
    Retorna dict: { comp_codigo: [ {grado_nombre, cap_orden, descripcion}, ... ] }
    Solo incluye filas con Estado == 'OK'.
    """
    if not os.path.exists(EXCEL_DESEMPENOS):
        print(f"   ⚠️ No encontrado archivo Excel: {EXCEL_DESEMPENOS}")
        return {}

    wb = openpyxl.load_workbook(EXCEL_DESEMPENOS)
    result = {}

    for comp_codigo in range(1, 5):
        sheet_name = f"C{comp_codigo}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        desempenos = []
        for row in ws.iter_rows(values_only=True):
            if not row or row[4] != "OK":
                continue
            grado_nombre = str(row[0]).strip().upper() if row[0] else None
            cap_orden = int(row[1]) if row[1] else None
            descripcion = str(row[3]).strip() if row[3] else None
            if grado_nombre and cap_orden and descripcion:
                desempenos.append({
                    "grado_nombre": grado_nombre,
                    "cap_orden": cap_orden,
                    "descripcion": descripcion
                })
        result[comp_codigo] = desempenos
        print(f"   📊 {sheet_name}: {len(desempenos)} desempeños leídos del Excel")

    return result


def load_matematica_data():
    """Carga todos los datos de Matemática en la base de datos."""
    print("=" * 60)
    print("🔢 CARGADOR DE COMPETENCIAS MATEMÁTICAS")
    print("=" * 60)
    
    # Verificar carpeta de archivos Word
    if not os.path.exists(DOCX_FOLDER):
        print(f"❌ Error: No se encontró la carpeta: {DOCX_FOLDER}")
        return
    
    print(f"📂 Carpeta fuente: {DOCX_FOLDER}")
    
    # Crear tablas si no existen
    Base.metadata.create_all(bind=engine)
    
    db = SessionLocal()
    
    try:
        # 1. Limpiar datos existentes de Matemática
        print("\n🗑️ Limpiando datos existentes de Matemática...")
        db.query(DesempenoMatematica).delete()
        db.query(EstandarMatematica).delete()
        db.query(CapacidadMatematica).delete()
        db.query(CompetenciaMatematica).delete()
        db.commit()
        
        # 2. Crear/Actualizar Grados (solo los de matemática si no existen)
        print("\n📚 Verificando grados...")
        grados_db = {}
        
        for grado_info in GRADOS_MATEMATICA:
            # Buscar si ya existe el grado
            existing = db.query(Grado).filter(
                Grado.nombre == grado_info["nombre"]
            ).first()
            
            if existing:
                grados_db[grado_info["nombre"]] = existing
            else:
                nuevo_grado = Grado(
                    nombre=grado_info["nombre"],
                    numero=grado_info["numero"],
                    nivel=grado_info["nivel"],
                    orden=grado_info["orden"]
                )
                db.add(nuevo_grado)
                db.flush()
                grados_db[grado_info["nombre"]] = nuevo_grado
                print(f"   ➕ Nuevo grado: {grado_info['nombre']}")
        
        db.commit()
        
        # 3. Crear Competencias y Capacidades
        print("\n🎯 Creando competencias y capacidades...")
        competencias_db = {}
        capacidades_db = {}
        
        for comp_data in COMPETENCIAS:
            # Crear competencia
            competencia = CompetenciaMatematica(
                codigo=comp_data["codigo"],
                nombre=comp_data["nombre"],
                descripcion=comp_data["descripcion"]
            )
            db.add(competencia)
            db.flush()
            competencias_db[comp_data["codigo"]] = competencia
            print(f"   ✓ Competencia {comp_data['codigo']}: {comp_data['nombre']}")
            
            # Crear capacidades de esta competencia
            for idx, cap_nombre in enumerate(comp_data["capacidades"], 1):
                capacidad = CapacidadMatematica(
                    orden=idx,
                    nombre=cap_nombre,
                    competencia_id=competencia.id
                )
                db.add(capacidad)
                db.flush()
                # Key: (competencia_codigo, orden_capacidad)
                capacidades_db[(comp_data["codigo"], idx)] = capacidad
        
        db.commit()
        
        # 4. Cargar desempeños desde Excel
        print("\n📊 Cargando desempeños desde Excel...")
        excel_data = load_desempenos_from_excel()

        desempeno_contador = {}
        for comp_codigo, desempenos in excel_data.items():
            competencia = competencias_db[comp_codigo]
            for desemp in desempenos:
                grado_nombre = desemp["grado_nombre"]
                cap_orden = desemp["cap_orden"]
                cap_key = (comp_codigo, cap_orden)

                if grado_nombre not in grados_db:
                    continue
                if cap_key not in capacidades_db:
                    continue

                contador_key = (grado_nombre, cap_key)
                desempeno_contador[contador_key] = desempeno_contador.get(contador_key, 0) + 1
                codigo = str(desempeno_contador[contador_key]).zfill(2)

                desempeno = DesempenoMatematica(
                    codigo=codigo,
                    descripcion=desemp["descripcion"],
                    grado_id=grados_db[grado_nombre].id,
                    capacidad_id=capacidades_db[cap_key].id
                )
                db.add(desempeno)

        db.commit()
        
        # 5. Estadísticas finales
        print("\n" + "=" * 60)
        print("✅ CARGA COMPLETADA")
        print("=" * 60)
        
        total_grados = db.query(Grado).count()
        total_competencias = db.query(CompetenciaMatematica).count()
        total_capacidades = db.query(CapacidadMatematica).count()
        total_estandares = db.query(EstandarMatematica).count()
        total_desempenos = db.query(DesempenoMatematica).count()
        
        print(f"\n📊 Resumen de datos cargados:")
        print(f"   - Grados totales (con Comunicación): {total_grados}")
        print(f"   - Competencias Matemáticas: {total_competencias}")
        print(f"   - Capacidades Matemáticas: {total_capacidades}")
        print(f"   - Estándares Matemáticos: {total_estandares}")
        print(f"   - Desempeños Matemáticos: {total_desempenos}")
        
        # Detalle por competencia
        print(f"\n📈 Desempeños por competencia:")
        for comp in db.query(CompetenciaMatematica).order_by(CompetenciaMatematica.codigo).all():
            count = db.query(DesempenoMatematica).join(CapacidadMatematica).filter(
                CapacidadMatematica.competencia_id == comp.id
            ).count()
            print(f"   {comp.codigo}. {comp.nombre[:45]}...: {count}")
        
    except Exception as e:
        db.rollback()
        print(f"\n❌ Error durante la carga: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    load_matematica_data()

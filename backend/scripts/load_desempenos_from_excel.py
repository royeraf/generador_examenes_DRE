"""
Script para cargar los desempeños corregidos de Matemática desde el archivo Excel
en la base de datos de producción (PostgreSQL en Render).

Ejecutar desde el directorio backend:
    python -m scripts.load_desempenos_from_excel

Archivo fuente:
    capacidades_mat/desempenos_extraidos.xlsx

¿Qué hace este script?
    1. Lee los desempeños desde las hojas C1, C2, C3, C4 del Excel
    2. Conecta a la base de datos de producción (PostgreSQL)
    3. Obtiene los IDs de grados y capacidades existentes en la BD
    4. Elimina SOLO los desempeños matemáticos existentes
    5. Inserta los desempeños corregidos del Excel (solo los que tienen estado "OK")
    6. Genera códigos secuenciales (01, 02, ...) por capacidad+grado

Esquema de base de datos relevante:
    - grados (id, nombre, numero, nivel, orden)
    - competencias_matematica (id, codigo, nombre, descripcion)
    - capacidades_matematica (id, orden, nombre, competencia_id)
    - desempenos_matematica (id, codigo, descripcion, grado_id, capacidad_id)

    Relaciones:
    - DesempenoMatematica.grado_id → Grado.id
    - DesempenoMatematica.capacidad_id → CapacidadMatematica.id
    - CapacidadMatematica.competencia_id → CompetenciaMatematica.id
    - CompetenciaMatematica.codigo = 1,2,3,4 (mapea a hojas C1,C2,C3,C4)
    - CapacidadMatematica.orden = 1,2,3,4 (mapea a columna "Cap #" del Excel)
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from app.core.database import Base
from app.models.db_models import (
    Grado,
    CompetenciaMatematica,
    CapacidadMatematica,
    DesempenoMatematica,
)
# Import all models so Base.metadata knows about them
from app.models.docente import Docente  # noqa: F401

# =====================================================
# CONFIGURACIÓN DE BASE DE DATOS
# =====================================================
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./desempenos.db")

# Convertir URL async a sync para SQLAlchemy síncrono
if "asyncpg" in DATABASE_URL:
    DATABASE_URL = DATABASE_URL.replace("postgresql+asyncpg://", "postgresql://")

if DATABASE_URL.startswith("postgresql"):
    engine = create_engine(DATABASE_URL)
else:
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})

SessionLocal = sessionmaker(bind=engine)

# =====================================================
# RUTA AL ARCHIVO EXCEL
# =====================================================
EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "capacidades_mat",
    "desempenos_extraidos.xlsx",
)

# Mapeo: hoja del Excel → código de competencia
SHEET_TO_COMPETENCIA = {
    "C1": 1,
    "C2": 2,
    "C3": 3,
    "C4": 4,
}


def load_desempenos_from_excel():
    """Carga los desempeños corregidos desde el Excel a la base de datos."""
    print("=" * 60)
    print("📊 CARGA DE DESEMPEÑOS DESDE EXCEL")
    print("=" * 60)

    # Verificar que el Excel existe
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ No se encontró el archivo: {EXCEL_PATH}")
        return

    print(f"📄 Archivo fuente: {EXCEL_PATH}")
    print(f"🔗 Base de datos: {DATABASE_URL[:50]}...")

    # Leer Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)

    db = SessionLocal()

    try:
        # 1. Obtener grados existentes (por nombre)
        grados_db = {g.nombre: g for g in db.query(Grado).all()}
        print(f"\n📚 Grados en BD: {len(grados_db)}")
        for nombre in sorted(grados_db.keys(), key=lambda n: grados_db[n].orden):
            print(f"   - {nombre} (id={grados_db[nombre].id})")

        # 2. Obtener competencias y capacidades existentes
        competencias_db = {
            c.codigo: c for c in db.query(CompetenciaMatematica).all()
        }
        print(f"\n🎯 Competencias en BD: {len(competencias_db)}")

        if not competencias_db:
            print("❌ No hay competencias en la BD. Ejecuta primero load_matematica.py")
            return

        # Mapeo: (competencia_codigo, orden_capacidad) → CapacidadMatematica
        capacidades_db = {}
        for cap in db.query(CapacidadMatematica).all():
            comp = competencias_db.get(None)
            # Buscar el código de la competencia
            for codigo, comp_obj in competencias_db.items():
                if comp_obj.id == cap.competencia_id:
                    capacidades_db[(codigo, cap.orden)] = cap
                    break

        print(f"📐 Capacidades en BD: {len(capacidades_db)}")
        for (comp_cod, cap_ord), cap in sorted(capacidades_db.items()):
            print(f"   - C{comp_cod} Cap {cap_ord}: {cap.nombre[:50]}... (id={cap.id})")

        # 3. Leer desempeños del Excel
        print("\n📖 Leyendo desempeños del Excel...")
        desempenos_excel = []

        for sheet_name, comp_codigo in SHEET_TO_COMPETENCIA.items():
            ws = wb[sheet_name]
            count = 0
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
                grado_nombre, cap_num, cap_name, descripcion, estado = row

                # Solo cargar los que tienen estado "OK"
                if estado != "OK":
                    continue

                if not grado_nombre or not cap_num or not descripcion:
                    continue

                cap_num = int(cap_num)

                desempenos_excel.append({
                    "grado_nombre": grado_nombre.strip(),
                    "competencia_codigo": comp_codigo,
                    "capacidad_orden": cap_num,
                    "descripcion": descripcion.strip(),
                })
                count += 1

            print(f"   {sheet_name} (Competencia {comp_codigo}): {count} desempeños")

        print(f"\n   Total desempeños a cargar: {len(desempenos_excel)}")

        # 4. Eliminar desempeños existentes
        existing_count = db.query(DesempenoMatematica).count()
        print(f"\n🗑️  Eliminando {existing_count} desempeños existentes...")
        db.query(DesempenoMatematica).delete()
        db.flush()

        # 5. Insertar nuevos desempeños
        print("📥 Insertando desempeños corregidos...")
        inserted = 0
        skipped = 0
        errors = []

        # Contador para códigos secuenciales: (grado_nombre, comp_codigo, cap_orden) → count
        codigo_counter = {}

        for desemp in desempenos_excel:
            grado_nombre = desemp["grado_nombre"]
            comp_cod = desemp["competencia_codigo"]
            cap_ord = desemp["capacidad_orden"]

            # Verificar que el grado existe
            if grado_nombre not in grados_db:
                errors.append(f"Grado no encontrado: '{grado_nombre}'")
                skipped += 1
                continue

            # Verificar que la capacidad existe
            cap_key = (comp_cod, cap_ord)
            if cap_key not in capacidades_db:
                errors.append(
                    f"Capacidad no encontrada: C{comp_cod} Cap {cap_ord}"
                )
                skipped += 1
                continue

            # Generar código secuencial
            counter_key = (grado_nombre, comp_cod, cap_ord)
            if counter_key not in codigo_counter:
                codigo_counter[counter_key] = 0
            codigo_counter[counter_key] += 1
            codigo = str(codigo_counter[counter_key]).zfill(2)

            nuevo = DesempenoMatematica(
                codigo=codigo,
                descripcion=desemp["descripcion"],
                grado_id=grados_db[grado_nombre].id,
                capacidad_id=capacidades_db[cap_key].id,
            )
            db.add(nuevo)
            inserted += 1

        db.commit()

        # 6. Estadísticas finales
        print("\n" + "=" * 60)
        print("✅ CARGA COMPLETADA")
        print("=" * 60)
        print(f"\n📊 Resultados:")
        print(f"   - Desempeños insertados: {inserted}")
        print(f"   - Desempeños omitidos: {skipped}")

        if errors:
            unique_errors = sorted(set(errors))
            print(f"\n⚠️  Errores ({len(unique_errors)} únicos):")
            for err in unique_errors:
                print(f"   - {err}")

        # Verificación: contar por competencia
        print(f"\n📈 Desempeños por competencia en BD:")
        for comp_codigo in sorted(competencias_db.keys()):
            comp = competencias_db[comp_codigo]
            count = (
                db.query(DesempenoMatematica)
                .join(CapacidadMatematica)
                .filter(CapacidadMatematica.competencia_id == comp.id)
                .count()
            )
            print(f"   C{comp.codigo}. {comp.nombre[:45]}...: {count}")

        # Verificación: contar por grado
        print(f"\n📈 Desempeños por grado en BD:")
        for grado_nombre in sorted(grados_db.keys(), key=lambda n: grados_db[n].orden):
            grado = grados_db[grado_nombre]
            count = (
                db.query(DesempenoMatematica)
                .filter(DesempenoMatematica.grado_id == grado.id)
                .count()
            )
            if count > 0:
                print(f"   {grado.nombre}: {count}")

        total_final = db.query(DesempenoMatematica).count()
        print(f"\n   Total desempeños en BD: {total_final}")

    except Exception as e:
        db.rollback()
        print(f"\n❌ Error durante la carga: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    load_desempenos_from_excel()
